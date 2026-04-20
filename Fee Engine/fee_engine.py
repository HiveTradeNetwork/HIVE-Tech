#!/usr/bin/env python3
"""
HIVE Fee Engine
===============
Calculates monthly service fees per Supply Partner and generates Excel invoices.

Run once per month after all supplier invoices for the period have been
entered in Airtable Transactions table.

Usage:
    python fee_engine.py "April 2026"

Requirements:
    pip install pyairtable openpyxl

Environment variables required:
    AIRTABLE_API_KEY  — your Airtable personal access token
"""

import os
import sys
from datetime import date, timedelta

from pyairtable import Api
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ─── CONFIGURATION ───────────────────────────────────────────────────────────

AIRTABLE_API_KEY = os.environ.get("AIRTABLE_API_KEY")

BASE_ID             = "appR2AC6MpDvVt7rc"
TABLE_TRANSACTIONS  = "tblxgZtg8QAoGb0v0"
TABLE_PARTNERS      = "tblB3wl3FwJr8l1Cg"
TABLE_FEE_CALCS     = "tblgkHfERzfOp08Ur"

FEE_RATE  = 0.02   # 2% service fee on member transaction value ex GST
GST_RATE  = 0.10   # 10% GST applied to HIVE's service fee

HIVE = {
    "name":          "HIVE Trade Network Pty Ltd",
    "abn":           "21 683 196 018",
    "email":         "admin@hivetradenetwork.com",
    "address":       "Central Coast NSW",
    "bank_name":     "Commonwealth Bank of Australia",
    "bsb":           "062-000",
    "account":       "12345678",
    "payment_terms": "14 days",
}

OUTPUT_DIR = "invoices"


# ─── AIRTABLE ────────────────────────────────────────────────────────────────

def get_api():
    if not AIRTABLE_API_KEY:
        sys.exit("ERROR: AIRTABLE_API_KEY environment variable not set.\n"
                 "Set it with: set AIRTABLE_API_KEY=your_token_here  (Windows)\n"
                 "             export AIRTABLE_API_KEY=your_token_here  (Mac/Linux)")
    return Api(AIRTABLE_API_KEY)


def fetch_transactions(api, fee_period):
    """Return all transaction records for the given fee period."""
    table = api.table(BASE_ID, TABLE_TRANSACTIONS)
    return table.all(formula=f"{{Fee Period}} = '{fee_period}'")


def fetch_partners(api):
    """Return all active Supply Partner records keyed by record ID."""
    table = api.table(BASE_ID, TABLE_PARTNERS)
    records = table.all(formula="{Status} = 'Active'")
    return {r["id"]: r["fields"] for r in records}


def write_fee_calculation(api, partner_id, data, invoice_number, fee_period, invoice_date, due_date):
    """Create a Fee Calculation record in Airtable."""
    table = api.table(BASE_ID, TABLE_FEE_CALCS)
    table.create({
        "Fee Period":                    fee_period,
        "Supply Partner":                [partner_id],
        "Total Transaction Value (AUD)": round(data["total_ex_gst"], 2),
        "Fee Rate Applied":              FEE_RATE,
        "Fee Amount (AUD)":              round(data["fee_amount"], 2),
        "Invoice Number":                invoice_number,
        "Invoice Date":                  invoice_date.strftime("%Y-%m-%d"),
        "Invoice Due Date":              due_date.strftime("%Y-%m-%d"),
        "Invoice Status":                "Draft",
        "Excel Invoice Generated":       False,
    })


# ─── FEE CALCULATION ─────────────────────────────────────────────────────────

def calculate_fees(transactions, partners):
    """
    Group transactions by Supply Partner and calculate 2% fee.
    Returns dict keyed by partner Airtable record ID.
    """
    totals = {}

    for record in transactions:
        f = record["fields"]
        partner_links = f.get("Supply Partner", [])
        amount_ex_gst = f.get("Amount Ex GST") or 0

        for pid in partner_links:
            if pid not in totals:
                p = partners.get(pid, {})
                totals[pid] = {
                    "partner_name":    p.get("Partner Name", "Unknown Partner"),
                    "billing_contact": p.get("Billing Contact Name", ""),
                    "billing_email":   p.get("Billing Email", ""),
                    "abn":             p.get("ABN", ""),
                    "total_ex_gst":    0,
                    "tx_count":        0,
                }
            totals[pid]["total_ex_gst"] += amount_ex_gst
            totals[pid]["tx_count"] += 1

    for data in totals.values():
        data["fee_amount"]   = round(data["total_ex_gst"] * FEE_RATE, 2)
        data["gst_on_fee"]   = round(data["fee_amount"] * GST_RATE, 2)
        data["total_payable"]= round(data["fee_amount"] + data["gst_on_fee"], 2)

    return totals


# ─── EXCEL INVOICE GENERATION ────────────────────────────────────────────────

GREEN  = "0B4550"
YELLOW = "E6FF2B"
WHITE  = "FFFFFF"
GREY   = "898A80"
LGREY  = "F9F7F2"


def _cell(ws, ref, value, bold=False, size=11, color=None, bg=None, italic=False, align="left", num_format=None):
    c = ws[ref]
    c.value = value
    c.font = Font(bold=bold, size=size, color=color or "000000", italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_format:
        c.number_format = num_format
    return c


def generate_invoice(data, fee_period, invoice_number, invoice_date, due_date, output_dir):
    """Generate a branded Excel tax invoice for one Supply Partner."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tax Invoice"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 22

    # ── Header band ──
    ws.merge_cells("A1:C1")
    _cell(ws, "A1", HIVE["name"], bold=True, size=15, color=WHITE, bg=GREEN)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:C2")
    _cell(ws, "A2",
          f"ABN {HIVE['abn']}   |   {HIVE['email']}   |   {HIVE['address']}",
          size=9, color=WHITE, bg=GREEN)
    ws.row_dimensions[2].height = 18

    # ── TAX INVOICE title ──
    ws.merge_cells("A4:C4")
    _cell(ws, "A4", "TAX INVOICE", bold=True, size=13)
    ws.row_dimensions[4].height = 22

    # ── Invoice meta ──
    for row, label, value, is_bold in [
        (6,  "Invoice Number", invoice_number,               True),
        (7,  "Invoice Date",   invoice_date.strftime("%d/%m/%Y"), False),
        (8,  "Payment Due",    due_date.strftime("%d/%m/%Y"),     True),
        (9,  "Fee Period",     fee_period,                   False),
    ]:
        _cell(ws, f"A{row}", label, color=GREY, size=10)
        _cell(ws, f"B{row}", value, bold=is_bold)

    # ── Bill To ──
    _cell(ws, "A11", "Bill To", bold=True)
    _cell(ws, "B11", data["partner_name"], bold=True)
    if data["billing_contact"]:
        _cell(ws, "A12", "Attention", color=GREY, size=10)
        _cell(ws, "B12", data["billing_contact"])
    if data["abn"]:
        _cell(ws, "A13", "ABN", color=GREY, size=10)
        _cell(ws, "B13", data["abn"])
    if data["billing_email"]:
        _cell(ws, "A14", "Email", color=GREY, size=10)
        _cell(ws, "B14", data["billing_email"])

    # ── Line items header ──
    ws.row_dimensions[16].height = 20
    for col, label in [("A", "Description"), ("B", "Amount (ex GST)"), ("C", "")]:
        _cell(ws, f"{col}16", label, bold=True, color=WHITE, bg=GREEN, size=10)

    # ── Line item ──
    ws["A17"] = f"HIVE member transaction service fee — {fee_period}"
    ws["B17"].value = data["fee_amount"]
    ws["B17"].number_format = '"$"#,##0.00'
    ws["C17"].value = (f"{data['tx_count']} member transactions  |  "
                       f"total member spend ${data['total_ex_gst']:,.2f} ex GST")
    ws["C17"].font = Font(size=9, color=GREY, italic=True)

    # ── Totals ──
    for row, label, value, bold, bg in [
        (19, "Subtotal (ex GST)",    data["fee_amount"],    False, None),
        (20, "GST (10%)",            data["gst_on_fee"],    False, None),
        (21, "TOTAL PAYABLE (AUD)",  data["total_payable"], True,  YELLOW),
    ]:
        _cell(ws, f"A{row}", label, bold=bold)
        c = ws[f"B{row}"]
        c.value = value
        c.font = Font(bold=bold)
        c.number_format = '"$"#,##0.00'
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)

    # ── Payment details ──
    _cell(ws, "A23", "Payment Details", bold=True)
    for row, label, value in [
        (24, "Bank",      HIVE["bank_name"]),
        (25, "BSB",       HIVE["bsb"]),
        (26, "Account",   HIVE["account"]),
        (27, "Reference", invoice_number),
    ]:
        _cell(ws, f"A{row}", label, color=GREY, size=10)
        _cell(ws, f"B{row}", value)

    ws.merge_cells("A29:C29")
    _cell(ws, "A29",
          f"Payment terms: {HIVE['payment_terms']} from invoice date. "
          "Please reference invoice number when transferring. Thank you.",
          italic=True, size=9, color=GREY)

    # ── Save ──
    os.makedirs(output_dir, exist_ok=True)
    safe_name = data["partner_name"].replace(" ", "_").replace("/", "-")
    path = os.path.join(output_dir, f"HIVE_Invoice_{invoice_number}_{safe_name}.xlsx")
    wb.save(path)
    return path


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        sys.exit('Usage: python fee_engine.py "April 2026"')

    fee_period  = sys.argv[1]
    today       = date.today()
    due_date    = today + timedelta(days=14)

    print(f"\nHIVE Fee Engine")
    print(f"Fee period : {fee_period}")
    print(f"Run date   : {today.strftime('%d/%m/%Y')}")
    print(f"Due date   : {due_date.strftime('%d/%m/%Y')}")
    print("=" * 50)

    api = get_api()

    print(f"Fetching transactions for '{fee_period}'...")
    transactions = fetch_transactions(api, fee_period)
    print(f"  Found {len(transactions)} transaction(s).")

    if not transactions:
        sys.exit("No transactions found for this period. "
                 "Check that Fee Period is set correctly on transaction records.")

    print("Fetching active Supply Partners...")
    partners = fetch_partners(api)
    print(f"  Found {len(partners)} active partner(s).")

    print("Calculating fees...")
    totals = calculate_fees(transactions, partners)

    print(f"\nResults for {fee_period}:")
    print("-" * 50)

    for i, (partner_id, data) in enumerate(totals.items(), start=1):
        invoice_number = f"HIVE-{today.strftime('%Y%m')}-{i:03d}"

        print(f"\n  {data['partner_name']}")
        print(f"    Transactions : {data['tx_count']}")
        print(f"    Total ex GST : ${data['total_ex_gst']:>12,.2f}")
        print(f"    Fee (2%)     : ${data['fee_amount']:>12,.2f}")
        print(f"    GST on fee   : ${data['gst_on_fee']:>12,.2f}")
        print(f"    Total due    : ${data['total_payable']:>12,.2f}")
        print(f"    Invoice no.  : {invoice_number}")

        write_fee_calculation(api, partner_id, data, invoice_number,
                              fee_period, today, due_date)

        path = generate_invoice(data, fee_period, invoice_number,
                                today, due_date, OUTPUT_DIR)
        print(f"    Saved        : {path}")

    print(f"\n{'=' * 50}")
    print(f"Done. {len(totals)} invoice(s) generated in /{OUTPUT_DIR}/")
    print("ACTION: Add bank details to HIVE dict in fee_engine.py before sending.")


if __name__ == "__main__":
    main()
